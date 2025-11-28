import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import (
    PieChart,
    BarChart,
    Reference,
    Series,
)
from openpyxl.chart.label import DataLabelList

def create_dashboard(wb, df):
    """
    Creates a Dashboard sheet with summary statistics and charts.
    """
    print("Creating Dashboard sheet...")
    # Create Dashboard sheet and move it to the front
    ws = wb.create_sheet("Dashboard", 0)
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    
    # --- Data Preparation ---
    
    # 1. Partners by Tier
    tier_counts = df['Partner Tier'].value_counts()
    
    # 2. Top 10 Countries
    country_counts = df['Country'].value_counts().head(10)
    
    # 3. Reference Coverage
    # Ensure 'Client References Count' is numeric, coerce errors to NaN then fill with 0
    df['Client References Count'] = pd.to_numeric(df['Client References Count'], errors='coerce').fillna(0)
    has_refs = (df['Client References Count'] > 0).sum()
    no_refs = len(df) - has_refs
    ref_stats = pd.Series({'With References': has_refs, 'No References': no_refs})

    # --- Writing Data Tables (Hidden/Side) ---
    
    # Helper to write a small table
    def write_table(start_row, start_col, title, data_series):
        # Title
        cell = ws.cell(row=start_row, column=start_col, value=title)
        cell.font = Font(bold=True, size=14)
        
        # Header
        ws.cell(row=start_row+1, column=start_col, value="Category").font = Font(bold=True)
        ws.cell(row=start_row+1, column=start_col+1, value="Count").font = Font(bold=True)
        
        # Data
        for i, (index, value) in enumerate(data_series.items()):
            ws.cell(row=start_row+2+i, column=start_col, value=str(index))
            ws.cell(row=start_row+2+i, column=start_col+1, value=value)
            
        return start_row + 2 + len(data_series)

    # Write tables starting at column A (we'll hide or just place charts over/next to them)
    # Actually, let's put data in columns A-B, D-E, G-H and charts below or next to them.
    # Better: Data in columns A-B, charts in C+.
    
    # Tier Data (A1)
    write_table(1, 1, "Partners by Tier", tier_counts)
    
    # Country Data (A10)
    write_table(10, 1, "Top 10 Countries", country_counts)
    
    # Reference Data (A25)
    write_table(25, 1, "Reference Coverage", ref_stats)
    
    # --- Creating Charts ---
    
    # 1. Pie Chart: Partners by Tier
    pie = PieChart()
    pie.title = "Partners by Tier"
    labels = Reference(ws, min_col=1, min_row=3, max_row=3+len(tier_counts)-1)
    data = Reference(ws, min_col=2, min_row=2, max_row=3+len(tier_counts)-1)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    ws.add_chart(pie, "D2")
    
    # 2. Bar Chart: Top 10 Countries
    bar = BarChart()
    bar.type = "col"
    bar.style = 10
    bar.title = "Top 10 Countries"
    bar.y_axis.title = 'Partners'
    bar.x_axis.title = 'Country'
    
    labels = Reference(ws, min_col=1, min_row=12, max_row=12+len(country_counts)-1)
    data = Reference(ws, min_col=2, min_row=11, max_row=12+len(country_counts)-1)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(labels)
    bar.legend = None
    ws.add_chart(bar, "D18")
    
    # 3. Donut Chart: Reference Coverage
    donut = PieChart()
    donut.title = "Partners with Client References"
    donut.doughnut = True
    
    labels = Reference(ws, min_col=1, min_row=27, max_row=28)
    data = Reference(ws, min_col=2, min_row=26, max_row=28)
    donut.add_data(data, titles_from_data=True)
    donut.set_categories(labels)
    donut.dataLabels = DataLabelList()
    donut.dataLabels.showPercent = True
    ws.add_chart(donut, "L2")

    # Hide gridlines for cleaner look
    ws.sheet_view.showGridLines = False
    
    # Adjust column widths for data tables
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 10

def format_excel(input_csv, output_excel):
    """
    Reads a CSV, converts it to Excel, and applies professional formatting.
    """
    print(f"Reading CSV from {input_csv}...")
    try:
        df = pd.read_csv(input_csv)
    except FileNotFoundError:
        print(f"Error: File not found at {input_csv}")
        return

    # Create output directory if needed
    os.makedirs(os.path.dirname(output_excel), exist_ok=True)

    print(f"Writing raw data to {output_excel}...")
    # Write to Excel using pandas first
    df.to_excel(output_excel, index=False, sheet_name='Partners')

    # Load workbook for formatting
    wb = load_workbook(output_excel)
    
    # --- Create Dashboard ---
    create_dashboard(wb, df)
    
    # --- Format Data Sheet ---
    ws = wb['Partners']

    # Define Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Dark Blue
    
    # Alternating row colors
    row_fill_even = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid") # Light Grey
    row_fill_odd = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")   # White
    
    border_style = Side(border_style="thin", color="D9D9D9")
    border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)

    print("Applying styles and formatting to Partners sheet...")
    
    # 1. Format Header
    for col_num, cell in enumerate(ws[1], 1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # 2. Format Rows (Zebra Striping & Borders)
    # Optimization: Only format used range
    for row_num, row in enumerate(ws.iter_rows(min_row=2), 2):
        fill = row_fill_even if row_num % 2 == 0 else row_fill_odd
        for cell in row:
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    # 3. Auto-adjust Column Widths
    for col_num, column in enumerate(ws.columns, 1):
        max_length = 0
        column_letter = get_column_letter(col_num)
        
        # Check header length
        header_val = ws.cell(row=1, column=col_num).value
        if header_val:
            max_length = len(str(header_val))
        
        # Check first 100 rows for content length to estimate width (performance optimization)
        for i, cell in enumerate(column):
            if i > 100: break 
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50) # Cap width at 50
        ws.column_dimensions[column_letter].width = adjusted_width

    # 4. Freeze Header
    ws.freeze_panes = "A2"

    # 5. Add AutoFilter
    ws.auto_filter.ref = ws.dimensions

    print(f"Saving formatted file to {output_excel}...")
    wb.save(output_excel)
    print("Done!")

if __name__ == "__main__":
    input_file = "data/raw/global_partners.csv"
    output_file = "data/processed/global_partners_styled.xlsx"
    format_excel(input_file, output_file)
